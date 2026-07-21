"""
余额汇总 - 一键合并处理脚本（快手磁力版）

用法:
  python 一键合并余额汇总.py             # 处理今天的日期目录
  python 一键合并余额汇总.py 2026-07-05  # 处理指定日期目录
  也可以直接双击同目录下的“一键运行.command”（macOS）

功能：
  处理快手磁力（磁力引擎）平台余额汇总：
  1. 扫描 ZIP 内的 XLSX 文件，只读取第一个可用文件的表头作为固定字段
  2. 一次只读取和处理一个 XLSX，避免把所有数据放入内存
  3. 后续文件按固定字段处理，多出的字段忽略，缺失字段记录警告并填充为空
  4. 计算四项余额之和，筛选总余额不为 0 的数据
  5. 使用 write_only 模式逐行写入两个 XLSX 结果文件
  6. 按首次出现顺序去重账户 ID
  7. 记录 ZIP 扫描、ZIP 解压/Excel 解析、计算、写入和保存耗时

处理结果统一写入 日期目录/生成文件/磁力- 余额汇总/
执行日志（含失败原因、耗时明细和详细堆栈）写入
日期目录/失败文件与日志/失败日志_<时间戳>.txt

说明：
  - ZIP 内的 XLSX 是以流的方式读取的，日志中的“ZIP 解压 + Excel 解析”
    是一个合并耗时；如果要把物理解压和 Excel 解析完全拆开，需要先把 XLSX
    解压到磁盘，这会增加额外的磁盘读写，反而可能变慢。
  - 最终 XLSX 单个工作表最多保存 1,048,575 行数据（不含表头）。超过限制时
    直接失败，不自动拆分文件。
  - 账户 ID 去重使用 Python 集合，不依赖 SQLite 或其他外部数据库。

依赖: pip install pandas openpyxl python-calamine
"""

import os
import platform
import subprocess
import sys
import tempfile
import time
import traceback
import zipfile
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook


# Excel 单个工作表最多 1,048,576 行，其中 1 行是表头。
MAX_EXCEL_DATA_ROWS = 1_048_575

SUM_COLUMNS = ["现金可用余额", "前返可用余额", "后返可用余额", "信用可用余额"]
ACCOUNT_ID_COLUMN = "账户ID"
TOTAL_BALANCE_COLUMN = "总可用余额"


# 获取 exe 或脚本的真实路径
if getattr(sys, 'frozen', False):
    # PyInstaller 打包后的 exe
    EXE_DIR = Path(sys.executable).resolve().parent
else:
    # 直接运行 Python 脚本
    EXE_DIR = Path(__file__).resolve().parent

# exe 所在目录就是项目根目录，日期目录跟 exe 平级
PROJECT_ROOT = EXE_DIR


# ============================================================
# 公共工具函数
# ============================================================


def format_seconds(seconds):
    """格式化耗时，保留 3 位小数便于比较小步骤。"""
    return f"{seconds:.3f}秒"


def is_missing_value(value):
    """判断单元格是否为空，兼容 None、NaN、pd.NA。"""
    if value is None:
        return True
    try:
        result = pd.isna(value)
        return bool(result) if not hasattr(result, "__len__") else False
    except (TypeError, ValueError):
        return False


def excel_cell_value(value):
    """把 pandas/numpy 空值和标量转换为 openpyxl 可写入的值。"""
    if is_missing_value(value):
        return None
    if hasattr(value, "item"):
        try:
            return value.item()
        except (ValueError, TypeError):
            pass
    return value


def create_temp_path(directory: Path, final_name: str) -> Path:
    """在目标目录创建一个尚未写入内容的临时路径。"""
    suffix = ".xlsx" if final_name.endswith(".xlsx") else ".tmp"
    fd, temp_name = tempfile.mkstemp(
        prefix=f".{final_name}.",
        suffix=suffix,
        dir=str(directory),
    )
    os.close(fd)
    temp_path = Path(temp_name)
    temp_path.unlink(missing_ok=True)
    return temp_path


def cleanup_path(path):
    """尽力删除临时文件，不覆盖真正的处理异常。"""
    if path is None:
        return
    try:
        path.unlink(missing_ok=True)
    except FileNotFoundError:
        pass
    except Exception:
        pass


def close_write_only_workbook(workbook):
    """异常路径安全关闭 write_only 工作簿，先关闭工作表写入生成器。"""
    if workbook is None:
        return

    # Workbook.close() 只关闭归档文件；如果工作表的行生成器仍在运行，
    # 直接关闭归档会导致生成器收尾时访问已关闭文件。
    for worksheet in workbook.worksheets:
        try:
            worksheet.close()
        except Exception:
            # 异常路径只做尽力清理，不覆盖真正的处理异常。
            pass

    try:
        workbook.close()
    except Exception:
        pass


def read_excel_member(zf: zipfile.ZipFile, member_name: str, nrows=None):
    """从 ZIP 内读取一个 XLSX；耗时由调用方统计。"""
    read_kwargs = {
        "dtype": str,
        "engine": "calamine",
    }
    if nrows is not None:
        read_kwargs["nrows"] = nrows

    with zf.open(member_name, "r") as file_obj:
        return pd.read_excel(file_obj, **read_kwargs)


def collect_excel_members(raw_dir: Path):
    """收集 ZIP 内的 XLSX，并在收集阶段验证 ZIP 可以正常打开。"""
    zip_files = list(raw_dir.glob("*.zip"))
    if not zip_files:
        raise RuntimeError(f"在 {raw_dir} 下未找到压缩包文件（*.zip）")

    members = []
    for zip_path in zip_files:
        with zipfile.ZipFile(zip_path, "r") as zf:
            for file_info in zf.infolist():
                if (
                    file_info.filename.endswith(".xlsx")
                    and not file_info.filename.startswith("__MACOSX")
                ):
                    members.append((zip_path, file_info.filename))

    if not members:
        raise RuntimeError(f"在 {raw_dir} 下的压缩包中未找到 Excel 文件（*.xlsx）")
    return zip_files, members


def get_first_file_columns(members, warning_messages):
    """只读取第一个可用 XLSX 的表头，后续文件按固定字段处理。"""
    header_time = 0.0

    for index, (zip_path, member_name) in enumerate(members, start=1):
        read_start = time.perf_counter()
        try:
            with zipfile.ZipFile(zip_path, "r") as zf:
                header_df = read_excel_member(zf, member_name, nrows=0)
        except Exception as exc:
            warning_messages.append(
                f"读取固定字段来源表头失败（{index}/{len(members)}）"
                f" {zip_path.name} / {member_name}: {exc}"
            )
            continue
        finally:
            header_time += time.perf_counter() - read_start

        if len(header_df.columns) == 0:
            warning_messages.append(
                f"固定字段来源 Excel 表头为空，继续尝试下一个文件: "
                f"{zip_path.name} / {member_name}"
            )
            continue

        all_columns = list(header_df.columns)
        missing_columns = [
            column
            for column in [*SUM_COLUMNS, ACCOUNT_ID_COLUMN]
            if column not in all_columns
        ]
        if missing_columns:
            raise RuntimeError(
                f"固定字段来源文件缺少必需列: {', '.join(missing_columns)}；"
                f"文件: {zip_path.name} / {member_name}"
            )

        return all_columns, header_time

    raise RuntimeError("未读取到任何有效 Excel 表头")


def account_id_key(value):
    """生成账户 ID 去重键，保留空值和空字符串的区别。"""
    if is_missing_value(value):
        return ("missing", "")
    return ("value", str(value))


def publish_files(temp_to_final):
    """全部临时文件准备完成后，再替换正式文件。"""
    for temp_path, final_path in temp_to_final:
        os.replace(temp_path, final_path)


# ============================================================
# 平台：快手磁力（磁力引擎）- 余额处理
# ============================================================


def process_ciliyinqing(date_dir: Path) -> dict:
    platform_name = "磁力- 余额汇总"
    raw_dir = date_dir / "原始文件" / platform_name
    output_dir = date_dir / "生成文件" / platform_name

    if not raw_dir.is_dir():
        raise RuntimeError(f"未找到原始文件目录: {raw_dir}")
    output_dir.mkdir(parents=True, exist_ok=True)

    process_start = time.perf_counter()
    timings = {
        "查找并扫描 ZIP": 0.0,
        "读取表头（含 ZIP 解压 + Excel 解析）": 0.0,
        "打开 ZIP": 0.0,
        "读取数据（含 ZIP 解压 + Excel 解析）": 0.0,
        "余额计算与筛选": 0.0,
        "汇总 XLSX 行写入": 0.0,
        "账户 ID 去重与写入": 0.0,
        "汇总 XLSX 保存": 0.0,
        "账户 ID XLSX 保存": 0.0,
        "正式文件替换": 0.0,
    }
    warning_messages = []
    file_timings = []

    scan_start = time.perf_counter()
    try:
        zip_files, members = collect_excel_members(raw_dir)
    except Exception as exc:
        timings["查找并扫描 ZIP"] = time.perf_counter() - scan_start
        raise RuntimeError(f"扫描压缩包失败: {exc}") from exc
    timings["查找并扫描 ZIP"] = time.perf_counter() - scan_start

    print(
        f"   找到 {len(zip_files)} 个 ZIP，{len(members)} 个 Excel，"
        f"扫描耗时 {format_seconds(timings['查找并扫描 ZIP'])}"
    )

    all_columns, header_time = get_first_file_columns(
        members,
        warning_messages,
    )
    timings["读取表头（含 ZIP 解压 + Excel 解析）"] = header_time

    output_columns = list(all_columns)
    if TOTAL_BALANCE_COLUMN not in output_columns:
        output_columns.append(TOTAL_BALANCE_COLUMN)

    today_str = date_dir.name
    merged_filename = f"磁力-余额汇总_{today_str}.xlsx"
    ids_filename = f"磁力-广告主账户ID_{today_str}.xlsx"
    merged_path = output_dir / merged_filename
    ids_path = output_dir / ids_filename

    temp_merged_path = create_temp_path(output_dir, merged_filename)
    temp_ids_path = create_temp_path(output_dir, ids_filename)

    summary_workbook = None
    ids_workbook = None
    total_filtered_rows = 0
    unique_account_id_count = 0
    account_id_output_rows = 0
    processed_files = 0
    seen_account_ids = set()

    try:
        summary_workbook = Workbook(write_only=True)
        summary_sheet = summary_workbook.create_sheet("Sheet1")
        summary_sheet.append([excel_cell_value(column) for column in output_columns])

        ids_workbook = Workbook(write_only=True)
        ids_sheet = ids_workbook.create_sheet("Sheet1")
        ids_sheet.append(["广告主账户ID"])

        # 按 ZIP 和 ZIP 内文件的原始顺序处理，保持原脚本的行顺序。
        for zip_path in zip_files:
            zip_members = [
                member_name
                for current_zip, member_name in members
                if current_zip == zip_path
            ]
            if not zip_members:
                continue

            zip_open_start = time.perf_counter()
            try:
                zip_file = zipfile.ZipFile(zip_path, "r")
            except Exception as exc:
                timings["打开 ZIP"] += time.perf_counter() - zip_open_start
                warning_messages.append(f"打开压缩包 {zip_path.name} 失败: {exc}")
                continue

            timings["打开 ZIP"] += time.perf_counter() - zip_open_start
            with zip_file as zf:
                for member_name in zip_members:
                    processed_files += 1
                    print(
                        f"   处理 Excel {processed_files}/{len(members)}: "
                        f"{zip_path.name} / {member_name}"
                    )

                    file_read_start = time.perf_counter()
                    try:
                        data = read_excel_member(zf, member_name)
                    except Exception as exc:
                        file_read_time = time.perf_counter() - file_read_start
                        timings["读取数据（含 ZIP 解压 + Excel 解析）"] += file_read_time
                        warning_messages.append(
                            f"读取压缩包内文件 {zip_path.name} / {member_name} 失败: {exc}"
                        )
                        file_timings.append(
                            {
                                "文件": f"{zip_path.name} / {member_name}",
                                "读取": file_read_time,
                                "计算": 0.0,
                                "汇总写入": 0.0,
                                "账户ID写入": 0.0,
                                "保留行数": 0,
                                "状态": "失败",
                            }
                        )
                        continue

                    file_read_time = time.perf_counter() - file_read_start
                    timings["读取数据（含 ZIP 解压 + Excel 解析）"] += file_read_time

                    extra_columns = [
                        column for column in data.columns if column not in all_columns
                    ]
                    missing_columns = [
                        column for column in all_columns if column not in data.columns
                    ]
                    if extra_columns or missing_columns:
                        schema_parts = []
                        if missing_columns:
                            schema_parts.append(
                                f"缺少字段: {', '.join(map(str, missing_columns))}"
                            )
                        if extra_columns:
                            schema_parts.append(
                                f"新增字段将忽略: {', '.join(map(str, extra_columns))}"
                            )
                        warning_messages.append(
                            f"文件字段与第一个有效文件不一致 "
                            f"{zip_path.name} / {member_name}: "
                            + "；".join(schema_parts)
                        )

                    calculate_start = time.perf_counter()
                    # 后续文件统一按照第一个有效文件的字段处理；缺失字段填充 NaN，
                    # 多出的字段不写入最终汇总文件。
                    data = data.reindex(columns=all_columns)
                    for column in SUM_COLUMNS:
                        data[column] = pd.to_numeric(
                            data[column],
                            errors="coerce",
                        ).fillna(0)

                    data[TOTAL_BALANCE_COLUMN] = data[SUM_COLUMNS].sum(axis=1)
                    filtered_data = data.loc[
                        data[TOTAL_BALANCE_COLUMN] != 0,
                        output_columns,
                    ]
                    calculate_time = time.perf_counter() - calculate_start
                    timings["余额计算与筛选"] += calculate_time

                    write_start = time.perf_counter()
                    current_file_filtered_rows = len(filtered_data)
                    for row in filtered_data.itertuples(index=False, name=None):
                        if total_filtered_rows >= MAX_EXCEL_DATA_ROWS:
                            raise RuntimeError(
                                f"最终余额汇总数据超过 Excel 单个工作表上限 "
                                f"{MAX_EXCEL_DATA_ROWS} 行；按照当前要求不拆分文件"
                            )
                        summary_sheet.append(
                            [excel_cell_value(value) for value in row]
                        )
                        total_filtered_rows += 1
                    write_time = time.perf_counter() - write_start
                    timings["汇总 XLSX 行写入"] += write_time

                    account_id_start = time.perf_counter()
                    for account_id in filtered_data[ACCOUNT_ID_COLUMN].tolist():
                        key = account_id_key(account_id)
                        if key in seen_account_ids:
                            continue

                        if account_id_output_rows >= MAX_EXCEL_DATA_ROWS:
                            raise RuntimeError(
                                f"唯一账户 ID 数量超过 Excel 单个工作表上限 "
                                f"{MAX_EXCEL_DATA_ROWS} 行；按照当前要求不拆分文件"
                            )

                        seen_account_ids.add(key)
                        ids_sheet.append(
                            [None if is_missing_value(account_id) else str(account_id)]
                        )
                        account_id_output_rows += 1
                        if not is_missing_value(account_id):
                            unique_account_id_count += 1
                    account_id_time = time.perf_counter() - account_id_start
                    timings["账户 ID 去重与写入"] += account_id_time

                    file_timings.append(
                        {
                            "文件": f"{zip_path.name} / {member_name}",
                            "读取": file_read_time,
                            "计算": calculate_time,
                            "汇总写入": write_time,
                            "账户ID写入": account_id_time,
                            "保留行数": current_file_filtered_rows,
                            "状态": "成功",
                        }
                    )

                    print(
                        f"      保留 {current_file_filtered_rows} 行；"
                        f"读取 {format_seconds(file_read_time)}，"
                        f"计算 {format_seconds(calculate_time)}，"
                        f"汇总写入 {format_seconds(write_time)}，"
                        f"账户ID去重写入 {format_seconds(account_id_time)}"
                    )

                    del filtered_data
                    del data

        if total_filtered_rows == 0:
            raise RuntimeError("所有行总可用余额均为 0，无需生成文件")

        save_start = time.perf_counter()
        summary_workbook.save(temp_merged_path)
        summary_workbook.close()
        summary_workbook = None
        timings["汇总 XLSX 保存"] = time.perf_counter() - save_start

        save_start = time.perf_counter()
        ids_workbook.save(temp_ids_path)
        ids_workbook.close()
        ids_workbook = None
        timings["账户 ID XLSX 保存"] = time.perf_counter() - save_start

        publish_start = time.perf_counter()
        publish_files(
            [
                (temp_merged_path, merged_path),
                (temp_ids_path, ids_path),
            ]
        )
        temp_merged_path = None
        temp_ids_path = None
        timings["正式文件替换"] = time.perf_counter() - publish_start

        timings["总耗时"] = time.perf_counter() - process_start
        return {
            "合并文件": str(merged_path),
            "合并行数": total_filtered_rows,
            "账户ID文件": str(ids_path),
            "唯一账户ID数": unique_account_id_count,
            "警告": warning_messages if warning_messages else None,
            "耗时明细": timings,
            "文件耗时": file_timings,
        }
    finally:
        close_write_only_workbook(summary_workbook)
        close_write_only_workbook(ids_workbook)
        cleanup_path(temp_merged_path)
        cleanup_path(temp_ids_path)


# ============================================================
# 主流程
# ============================================================


def show_alert(title, message):
    """在 macOS 上显示通知。"""
    if platform.system() == "Darwin":
        safe_message = (
            str(message)
            .replace("\\", "\\\\")
            .replace('"', '\\"')
            .replace("\n", "\\n")
        )
        safe_title = str(title).replace("\\", "\\\\").replace('"', '\\"')
        subprocess.run(
            [
                "osascript",
                "-e",
                f'display dialog "{safe_message}" with title "{safe_title}" '
                'buttons {"确定"} default button 1 with icon caution',
            ],
            check=False,
        )
    else:
        print(f"{title}: {message}")


def get_target_date_str():
    """获取目标日期字符串，支持命令行传参，默认今天。"""
    if len(sys.argv) > 1:
        return sys.argv[1]
    return datetime.now().strftime("%Y-%m-%d")


def write_log(log_path: Path, date_str: str, results):
    """写入处理结果、各步骤耗时和失败日志。"""
    lines = [
        f"余额汇总处理日志 - {date_str}",
        f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "=" * 60,
        "",
    ]

    for platform_name, result, error in results:
        if result is not None:
            lines.extend(
                [
                    f"✅ {platform_name} - 成功",
                    f"  合并行数: {result['合并行数']}",
                    f"  唯一账户ID数: {result['唯一账户ID数']}",
                    f"  合并文件: {result['合并文件']}",
                    f"  账户ID文件: {result['账户ID文件']}",
                    "",
                    "  各步骤耗时:",
                ]
            )
            for name, seconds in result.get("耗时明细", {}).items():
                lines.append(f"    {name}: {format_seconds(seconds)}")

            if result.get("文件耗时"):
                lines.extend(["", "  各 Excel 文件耗时:"])
                for item in result["文件耗时"]:
                    lines.append(
                        f"    {item['状态']} {item['文件']} | "
                        f"读取 {format_seconds(item['读取'])} | "
                        f"计算 {format_seconds(item['计算'])} | "
                        f"汇总写入 {format_seconds(item['汇总写入'])} | "
                        f"账户ID写入 {format_seconds(item['账户ID写入'])} | "
                        f"保留 {item['保留行数']} 行"
                    )

            if result.get("警告"):
                lines.extend(["", "  警告信息:"])
                lines.extend(f"    • {warning}" for warning in result["警告"])
        else:
            lines.extend(
                [
                    f"❌ {platform_name} - 失败",
                    f"  失败原因: {error['error']}",
                    "  详细堆栈:",
                    *[f"    {line}" for line in error["traceback"].splitlines()],
                ]
            )
        lines.extend(["", "=" * 60, ""])

    log_path.write_text("\n".join(lines), encoding="utf-8")


def main():
    date_str = get_target_date_str()
    date_dir = PROJECT_ROOT / date_str

    if not date_dir.is_dir():
        message = (
            f"未找到日期目录: {date_dir}\n\n"
            "请先运行 创建日期目录.py 创建目录结构，再放入原始文件。"
        )
        print(f"❌ {message}")
        show_alert("余额汇总处理失败", message)
        sys.exit(1)

    log_dir = date_dir / "失败文件与日志"
    log_dir.mkdir(parents=True, exist_ok=True)

    tasks = [
        ("磁力- 余额汇总", process_ciliyinqing),
    ]

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = log_dir / f"失败日志_{timestamp}.txt"
    all_results = []

    for platform_name, func in tasks:
        try:
            print(f"🔄 开始处理 {platform_name}...")
            result = func(date_dir)
            all_results.append((platform_name, result, None))

            total_time = result["耗时明细"].get("总耗时", 0.0)
            print(f"✅ {platform_name} 处理完成 ({format_seconds(total_time)})")
            print(f"   合并行数: {result['合并行数']}")
            print(f"   唯一账户ID数: {result['唯一账户ID数']}")
            print(f"   合并文件: {result['合并文件']}")
            print(f"   账户ID文件: {result['账户ID文件']}")
            print("   各步骤耗时:")
            for name, seconds in result["耗时明细"].items():
                print(f"      {name}: {format_seconds(seconds)}")
        except Exception as exc:
            error_traceback = traceback.format_exc()
            all_results.append(
                (
                    platform_name,
                    None,
                    {"error": str(exc), "traceback": error_traceback},
                )
            )
            print(f"❌ {platform_name} 处理失败: {exc}")

    write_log(log_path, date_str, all_results)

    success_count = sum(1 for _, result, _ in all_results if result is not None)
    fail_count = len(all_results) - success_count

    if fail_count > 0:
        print(f"\n⚠️ 部分平台处理失败，详细日志请查看: {log_path}")
        show_alert(
            "余额汇总处理完成",
            f"处理完成，{success_count}个成功，{fail_count}个失败。\n"
            "详细日志请查看失败文件与日志目录。",
        )
    else:
        print(f"\n✅ 所有平台处理成功！详细日志已保存: {log_path}")
        show_alert("余额汇总处理成功", "所有平台余额汇总处理完成！")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n❌ 发生错误: {e}")
    finally:
        input("\n按回车键退出...")
