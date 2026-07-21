"""
总消耗汇总脚本

功能：
  读取所有平台生成文件中的回填结果，提取各项消耗数据，
  汇总到一个Excel文件中，包含各细分消耗列。

用法：
  python 总消耗汇总.py                    # 汇总今天的日期目录
  python 总消耗汇总.py 2026-07-08         # 汇总指定日期目录

依赖: pip install pandas openpyxl
"""

import sys
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl.utils import get_column_letter


# 获取 exe 或脚本的真实路径
if getattr(sys, 'frozen', False):
    # PyInstaller 打包后的 exe
    EXE_DIR = Path(sys.executable).resolve().parent
else:
    # 直接运行 Python 脚本
    EXE_DIR = Path(__file__).resolve().parent

# exe 所在目录就是项目根目录，日期目录跟 exe 平级
PROJECT_ROOT = EXE_DIR

# 各平台配置: (目录名, 平台显示名)
PLATFORM_CONFIG = [
    ("快手-磁力- 消耗", "快手-磁力"),
    ("快手-金教- 消耗", "快手-金教"),
    ("快手-本地生活- 消耗", "快手-本地生活"),
    ("金牛-华北- 消耗", "金牛-华北"),
    ("金牛-广分- 消耗", "金牛-广分"),
    ("金牛-澳比- 消耗", "金牛-澳比"),
    ("抖音-竞价- 消耗", "抖音-竞价"),
    ("抖音-品牌- 消耗", "抖音-品牌"),
    ("抖音-本地推- 消耗", "抖音-本地推"),
    ("抖音-巨懂车- 消耗", "抖音-巨懂车"),
    ("千川-竞价- 消耗", "千川-竞价"),
    ("千川-品牌- 消耗", "千川-品牌"),
    ("千川-巨懂车- 消耗", "千川-巨懂车"),
    ("陵致长风-巨量- 消耗", "陵致长风-巨量"),
    ("陵致长风-千川- 消耗", "陵致长风-千川"),
    ("广点通 ADQ- 消耗", "广点通ADQ"),
    ("支付宝- 消耗", "支付宝"),
    ("B站-信息流- 消耗", "B站-信息流"),
    ("B站-商业起飞- 消耗", "B站-商业起飞"),
]


def find_backfill_file(output_dir: Path):
    """查找回填结果文件"""
    patterns = ["*回填结果*.xlsx", "*回填*.xlsx"]
    for pattern in patterns:
        files = [f for f in output_dir.glob(pattern) if not f.name.startswith('~$')]
        if files:
            return files[0]
    return None


def get_target_date_str():
    if len(sys.argv) > 1:
        return sys.argv[1]
    return datetime.now().strftime("%Y-%m-%d")


def main():
    date_str = get_target_date_str()
    date_dir = PROJECT_ROOT / date_str

    if not date_dir.is_dir():
        print(f"❌ 未找到日期目录: {date_dir}")
        sys.exit(1)

    print(f"📅 日期: {date_str}")
    print(f"📂 目录: {date_dir}")
    print()

    gen_dir = date_dir / "生成文件"
    output_file = gen_dir / f"总消耗汇总_{date_str}.xlsx"
    summary_rows = []

    for folder_name, display_name in PLATFORM_CONFIG:
        output_dir = gen_dir / folder_name
        if not output_dir.is_dir():
            continue

        backfill_file = find_backfill_file(output_dir)
        if not backfill_file:
            continue

        try:
            df = pd.read_excel(backfill_file)

            # 找到总消耗列
            total_col = None
            for col in df.columns:
                if '总消耗' in str(col):
                    total_col = col
                    break

            if not total_col:
                continue

            # 转换所有数值列
            numeric_cols = []
            for col in df.columns:
                if col.startswith('*日期') or 'ID' in col or '账户' in col or '备注' in col:
                    continue
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
                numeric_cols.append(col)

            # 求和得到各列汇总
            row = {"平台": display_name}
            for col in numeric_cols:
                row[col] = df[col].sum()

            summary_rows.append(row)
            print(f"  ✅ {display_name}: 总消耗 {row.get(total_col, 0):,.4f}")

        except Exception as e:
            print(f"  ⚠️  {display_name}: 读取失败 - {e}")

    if not summary_rows:
        print("❌ 没有找到任何回填数据")
        sys.exit(1)

    # 构造汇总DataFrame
    summary_df = pd.DataFrame(summary_rows).fillna(0)

    # 添加合计行
    totals = {"平台": "合计"}
    for col in summary_df.columns:
        if col != "平台":
            totals[col] = summary_df[col].sum()
    summary_df.loc[len(summary_df)] = totals

    # 写入Excel：汇总sheet + 每个平台单独一个sheet
    if output_file.exists():
        output_file.unlink()
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 汇总sheet
        summary_df.to_excel(writer, sheet_name="汇总", index=False)
        ws = writer.sheets["汇总"]
        for col_idx in range(2, len(summary_df.columns) + 1):
            for row_idx in range(2, len(summary_df) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.number_format = '0.000'
            ws.column_dimensions[get_column_letter(col_idx)].width = 16
        ws.column_dimensions['A'].width = 18

        # 每个平台单独一个sheet（写入回填结果的明细数据）
        for folder_name, display_name in PLATFORM_CONFIG:
            output_dir = gen_dir / folder_name
            if not output_dir.is_dir():
                continue
            backfill_file = find_backfill_file(output_dir)
            if not backfill_file:
                continue
            try:
                df = pd.read_excel(backfill_file)
                # sheet名最多31字符
                sheet_name = display_name[:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                # 设置数值列格式
                ws_platform = writer.sheets[sheet_name]
                for col_idx, col_name in enumerate(df.columns, 1):
                    if '日期' in str(col_name) or 'ID' in str(col_name) or '账户' in str(col_name) or '备注' in str(col_name):
                        continue
                    for row_idx in range(2, len(df) + 2):
                        cell = ws_platform.cell(row=row_idx, column=col_idx)
                        cell.number_format = '0.000'
            except Exception as e:
                print(f"  ⚠️  写入 {display_name} sheet失败: {e}")

    print()
    print(f"✅ 总消耗汇总已保存: {output_file}")
    print()

    # 打印汇总
    print("=" * 60)
    num_cols = [c for c in summary_df.columns if c != "平台"]
    header = f"{'平台':<14}" + "".join(f"{c:>14}" for c in num_cols)
    print(header)
    print("-" * 60)
    for _, row in summary_df.iterrows():
        line = f"{row['平台']:<14}"
        for c in num_cols:
            line += f"{row[c]:>14,.3f}"
        print(line)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n❌ 发生错误: {e}")
    finally:
        input("\n按回车键退出...")
